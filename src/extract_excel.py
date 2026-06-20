"""Extração segura do arquivo .xlsm sem alterar o original."""
import os
import logging
import hashlib
from pathlib import Path

import openpyxl
import pandas as pd
import yaml

logger = logging.getLogger(__name__)


def load_config(config_path: str = "config.yaml") -> dict:
    with open(config_path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def file_md5(path: str) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _find_sheet(wb: openpyxl.Workbook, name: str, fallback_index: int) -> openpyxl.worksheet.worksheet.Worksheet:
    """Localiza aba por nome parcial (case-insensitive) ou por índice."""
    name_lower = name.lower()
    for sname in wb.sheetnames:
        if sname.lower().startswith(name_lower[:20]):
            return wb[sname]
    # fallback por índice
    logger.warning("Aba '%s' não encontrada. Usando índice %d: '%s'", name, fallback_index, wb.sheetnames[fallback_index])
    return wb.worksheets[fallback_index]


def extract(config: dict) -> pd.DataFrame:
    excel_cfg = config["excel"]
    local_path = os.environ.get("EXCEL_LOCAL_PATH") or excel_cfg["local_path"]
    local_path = Path(local_path)

    if not local_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {local_path}")

    logger.info("Lendo: %s", local_path.name)
    md5 = file_md5(str(local_path))
    logger.info("MD5: %s", md5)

    wb = openpyxl.load_workbook(str(local_path), read_only=True, keep_vba=True, data_only=True)
    ws = _find_sheet(wb, excel_cfg["sheet_name"], excel_cfg.get("sheet_name_fallback_index", 0))
    logger.info("Aba selecionada: '%s' | max_row=%s", ws.title, ws.max_row)

    rows = ws.iter_rows(values_only=True)
    raw_header = list(next(rows))

    # Truncar nas últimas colunas com cabeçalho real (evita 16k colunas vazias do xlsm)
    last_col = 0
    for i, v in enumerate(raw_header):
        if v is not None and str(v).strip():
            last_col = i
    last_col += 1  # índice exclusivo

    raw_header = raw_header[:last_col]
    header = [str(c).strip() if c is not None else f"_col{i}" for i, c in enumerate(raw_header)]

    data = []
    for row in rows:
        row_slice = row[:last_col]
        if any(v is not None for v in row_slice):
            data.append(row_slice)

    wb.close()

    df = pd.DataFrame(data, columns=header)
    logger.info("Registros extraídos: %d", len(df))
    return df, md5, str(local_path.name)
