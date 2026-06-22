"""
Gera a PLANILHA DE CONFERÊNCIA (.xlsx) — arquivo separado, não toca no .xlsm.

Espelha a aba de dados original (32 colunas verbatim) e INSERE as colunas de
conferência AO LADO dos blocos originais:
  - bloco BARRAGEM  -> LAT/LON (CONFERIDA), STATUS CONF., DESLOC. (m)
  - bloco CASA FORÇA -> LAT/LON (CONFERIDA), STATUS CONF., DESLOC. (m)
  - meta no fim -> GRAU CONFIANÇA, FASE VERIFICADA, OBS, DATA, REVISADO

Os valores de conferência vêm do rastreador (conferencia_rastreador.csv).
Pode ser re-executado a cada lote revisado — sempre regenera o arquivo.
"""
import os, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
import yaml

sys.path.insert(0, os.path.dirname(__file__))
from geo_utils import safe_float, distancia_metros, STATUS_VALIDOS, GRAU_VALIDOS

SAIDA = "data/processed/conferencia_coordenadas.xlsx"
RASTREADOR = "data/processed/conferencia_rastreador.csv"

# Cores
AZUL = "1F4E78"; AZUL_CLARO = "DDEBF7"
VERDE = "375623"; VERDE_CLARO = "E2EFDA"   # casa de força (verde, como o círculo)
AMARELO = "806000"; AMARELO_CLARO = "FFF2CC"  # barragem (amarelo, como o círculo)
CINZA_CLARO = "F2F2F2"


CACHE_ORIGEM = "data/processed/origem_xlsm_cache.pkl"


def ler_xlsm_verbatim():
    """Lê as 32 colunas originais do .xlsm (read-only). Cacheia o resultado e,
    se o .xlsm tiver sido movido/fechado, usa o cache — não interrompe o fluxo."""
    import os, pickle
    with open("config.yaml", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    path = cfg["excel"]["local_path"]

    if not os.path.exists(path):
        if os.path.exists(CACHE_ORIGEM):
            with open(CACHE_ORIGEM, "rb") as fh:
                header, data = pickle.load(fh)
            print(f"AVISO: .xlsm não encontrado em {path}; usando cache {CACHE_ORIGEM}.")
            return header, data
        raise FileNotFoundError(
            f".xlsm não encontrado em {path} e sem cache em {CACHE_ORIGEM}. "
            "Ajuste excel.local_path no config.yaml.")

    wb = openpyxl.load_workbook(path, read_only=True, keep_vba=True, data_only=True)
    ws = next((wb[s] for s in wb.sheetnames
               if s.lower().startswith("central de processos hidrel")), wb.worksheets[0])
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    last = max((i for i, v in enumerate(header) if v not in (None, "")), default=0) + 1
    header = [str(h).strip() if h is not None else f"_c{i}" for i, h in enumerate(header[:last])]
    data = [list(r[:last]) for r in rows if any(v is not None for v in r[:last])]
    wb.close()

    with open(CACHE_ORIGEM, "wb") as fh:   # snapshot p/ resiliência
        pickle.dump((header, data), fh)
    return header, data


def main():
    header, data = ler_xlsm_verbatim()
    rast = pd.read_csv(RASTREADOR, encoding="utf-8-sig", dtype=str).fillna("")

    # localizar índices dos blocos
    def idx(nome):
        for i, h in enumerate(header):
            if h.strip().upper() == nome.upper():
                return i
        return None
    i_lonbar_kmz = idx("LON. BARRAGEM .KMZ")      # fim do bloco barragem (V)
    i_loncf_kmz = idx("LONGITUDE CASA FORÇA .KMZ")  # fim do bloco casa força (Z)

    # Definição das colunas novas
    novas_bar = ["LAT BARRAGEM (CONFERIDA)", "LON BARRAGEM (CONFERIDA)",
                 "STATUS CONF. BARRAGEM", "DESLOC. BARRAGEM (m)"]
    novas_cf = ["LAT CASA FORÇA (CONFERIDA)", "LON CASA FORÇA (CONFERIDA)",
                "STATUS CONF. CASA FORÇA", "DESLOC. CASA FORÇA (m)"]
    novas_meta = ["ANEEL FASE", "ANEEL LAT", "ANEEL LON", "ANEEL DIST (m)",
                  "GRAU CONFIANÇA", "FASE VERIFICADA", "OBS CONFERÊNCIA",
                  "DATA CONFERÊNCIA", "ORIGEM CONFERÊNCIA", "REVISADO"]

    # Monta a nova ordem de colunas (lista de tuplas: (origem, chave))
    # origem: 'orig' (idx no header original) | 'bar'/'cf'/'meta' (chave do rastreador)
    plano = []
    for i, h in enumerate(header):
        plano.append(("orig", i, h))
        if i == i_lonbar_kmz:
            for nb, key in zip(novas_bar, ["lat_barragem_conf", "lon_barragem_conf",
                                           "status_barragem_conf", "_dist_bar"]):
                plano.append(("bar", key, nb))
        if i == i_loncf_kmz:
            for nc, key in zip(novas_cf, ["lat_casaforca_conf", "lon_casaforca_conf",
                                          "status_casaforca_conf", "_dist_cf"]):
                plano.append(("cf", key, nc))
    for nm, key in zip(novas_meta, ["siga_fase", "siga_lat", "siga_lon", "siga_dist_m",
                                    "grau_confianca", "fase_verificada",
                                    "obs_conferencia", "data_conferencia",
                                    "origem_conferencia", "revisado"]):
        plano.append(("meta", key, nm))

    # Cria workbook de saída
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conferência Coordenadas"

    # Cabeçalho
    for c, (origem, key, nome) in enumerate(plano, start=1):
        cell = ws.cell(row=1, column=c, value=nome)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if origem == "orig":
            cell.fill = PatternFill("solid", fgColor=AZUL)
        elif origem == "bar":
            cell.fill = PatternFill("solid", fgColor=AMARELO)
        elif origem == "cf":
            cell.fill = PatternFill("solid", fgColor=VERDE)
        else:
            cell.fill = PatternFill("solid", fgColor=AZUL)

    # Linhas de dados
    for r_i, drow in enumerate(data):
        rast_row = rast.iloc[r_i] if r_i < len(rast) else None
        excel_row = r_i + 2
        for c, (origem, key, nome) in enumerate(plano, start=1):
            if origem == "orig":
                val = drow[key]
            elif origem in ("bar", "cf"):
                if key == "_dist_bar":
                    val = _dist(rast_row, "nav_lat_barragem", "nav_lon_barragem",
                                "lat_barragem_conf", "lon_barragem_conf")
                elif key == "_dist_cf":
                    val = _dist(rast_row, "nav_lat_casaforca", "nav_lon_casaforca",
                                "lat_casaforca_conf", "lon_casaforca_conf")
                else:
                    val = rast_row.get(key, "") if rast_row is not None else ""
            else:  # meta
                val = rast_row.get(key, "") if rast_row is not None else ""
            cell = ws.cell(row=excel_row, column=c, value=val if val != "" else None)
            cell.font = Font(size=9)
            # fundo leve para as colunas novas
            if origem == "bar":
                cell.fill = PatternFill("solid", fgColor=AMARELO_CLARO)
            elif origem == "cf":
                cell.fill = PatternFill("solid", fgColor=VERDE_CLARO)
            elif origem == "meta":
                cell.fill = PatternFill("solid", fgColor=CINZA_CLARO)

    # Larguras e freeze
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(plano))}1"
    for c, (origem, key, nome) in enumerate(plano, start=1):
        ws.column_dimensions[get_column_letter(c)].width = (
            34 if "OBS" in nome else 16 if origem != "orig" else 14)
    ws.row_dimensions[1].height = 42

    # Validação de dados (dropdowns) p/ colunas de status, grau, revisado
    def col_letter_de(nome):
        for c, (_, _, n) in enumerate(plano, start=1):
            if n == nome:
                return get_column_letter(c)
        return None
    nlin = len(data) + 1
    for nome, opcoes in [("STATUS CONF. BARRAGEM", STATUS_VALIDOS),
                         ("STATUS CONF. CASA FORÇA", STATUS_VALIDOS),
                         ("GRAU CONFIANÇA", GRAU_VALIDOS),
                         ("REVISADO", ["sim", "não"])]:
        cl = col_letter_de(nome)
        if cl:
            dv = DataValidation(type="list", formula1='"' + ",".join(opcoes) + '"',
                                allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{cl}2:{cl}{nlin}")

    # Aba de legenda
    leg = wb.create_sheet("Legenda")
    legendas = [
        ["PLANILHA DE CONFERÊNCIA GEOESPACIAL — HIDRELÉTRICAS IAT", ""],
        ["", ""],
        ["Esta planilha é uma CÓPIA de conferência.", ""],
        ["O arquivo .xlsm original NÃO foi modificado.", ""],
        ["", ""],
        ["Cor do cabeçalho", "Significado"],
        ["Azul", "Coluna original (verbatim do .xlsm)"],
        ["Amarelo", "Conferência da BARRAGEM (ponto central do barramento)"],
        ["Verde", "Conferência da CASA DE FORÇA (prédio com conduto forçado)"],
        ["", ""],
        ["STATUS possíveis", " / ".join(STATUS_VALIDOS)],
        ["GRAU DE CONFIANÇA", " / ".join(GRAU_VALIDOS)],
        ["DESLOC. (m)", "Distância entre coordenada original e a conferida (Haversine)"],
    ]
    for r_i, (a, b) in enumerate(legendas, start=1):
        leg.cell(row=r_i, column=1, value=a).font = Font(bold=(r_i in (1, 6)))
        leg.cell(row=r_i, column=2, value=b)
    leg.column_dimensions["A"].width = 40
    leg.column_dimensions["B"].width = 70

    os.makedirs("data/processed", exist_ok=True)
    wb.save(SAIDA)

    revisados = (rast["revisado"].str.lower() == "sim").sum() if "revisado" in rast else 0
    print(f"Planilha de conferência: {SAIDA}")
    print(f"  Colunas: {len(plano)} (originais {len(header)} + novas {len(plano) - len(header)})")
    print(f"  Linhas: {len(data)}")
    print(f"  Já revisados: {revisados}")


def _dist(rast_row, lat_o, lon_o, lat_c, lon_c):
    if rast_row is None:
        return None
    lo = safe_float(rast_row.get(lat_o, "")); loo = safe_float(rast_row.get(lon_o, ""))
    lc = safe_float(rast_row.get(lat_c, "")); lcc = safe_float(rast_row.get(lon_c, ""))
    if None in (lo, loo, lc, lcc):
        return None
    return round(distancia_metros(lo, loo, lc, lcc), 1)


if __name__ == "__main__":
    main()
