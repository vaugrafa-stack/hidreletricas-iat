"""Diagnóstico READ-ONLY da estrutura do .xlsm — não altera o original."""
import openpyxl
from openpyxl.utils import get_column_letter
import yaml

with open("config.yaml", encoding="utf-8") as f:
    cfg = yaml.safe_load(f)

path = cfg["excel"]["local_path"]
wb = openpyxl.load_workbook(path, read_only=True, keep_vba=True, data_only=True)

print("ABAS:", wb.sheetnames)
ws = None
for s in wb.sheetnames:
    if s.lower().startswith("central de processos hidrel"):
        ws = wb[s]
        break
if ws is None:
    ws = wb.worksheets[0]
print(f"\nABA: '{ws.title}' | max_row={ws.max_row} | max_col={ws.max_column}")

rows = ws.iter_rows(values_only=True)
header = list(next(rows))

# Truncar colunas vazias
last = 0
for i, v in enumerate(header):
    if v is not None and str(v).strip():
        last = i
last += 1
header = header[:last]

print(f"\n=== COLUNAS ({last}) ===")
for i, h in enumerate(header):
    letra = get_column_letter(i + 1)
    nome = str(h).strip() if h is not None else "(vazio)"
    print(f"  [{i:2d}] {letra:>3}  {nome}")

# Ler dados
data = []
for row in rows:
    rs = row[:last]
    if any(v is not None for v in rs):
        data.append(rs)
print(f"\nTotal de registros (linhas de dados): {len(data)}")

# Índices de colunas de interesse
def idx_of(nome):
    for i, h in enumerate(header):
        if h is not None and str(h).strip().upper() == nome.upper():
            return i
    return None

cols_interesse = ["TIPO", "SITUAÇÃO", "TIPOS DE LICENÇA",
                  "LATITUDE BARRAGEM", "LONGITUDE BARRAGEM",
                  "LAT. BARRAGEM .KMZ", "LON. BARRAGEM .KMZ",
                  "LATITUDE CASA FORÇA", "LONGITUDE CASA FORÇA",
                  "LAT. CASA FORÇA .KMZ", "LON. CASA FORÇA .KMZ",
                  "STATUS BARRAGEM", "STATUS CASA FORÇA"]
idxs = {c: idx_of(c) for c in cols_interesse}
print("\n=== ÍNDICES COLUNAS-CHAVE ===")
for c, i in idxs.items():
    letra = get_column_letter(i + 1) if i is not None else "?"
    print(f"  {c:28s} -> idx {i} ({letra})")

# Contagens
from collections import Counter
def col_vals(nome):
    i = idx_of(nome)
    if i is None:
        return []
    return [r[i] for r in data]

print("\n=== POR TIPO ===")
for k, v in Counter([str(x).strip() if x else "(vazio)" for x in col_vals("TIPO")]).most_common():
    print(f"  {k:20s} {v}")

print("\n=== POR SITUAÇÃO ===")
for k, v in Counter([str(x).strip() if x else "(vazio)" for x in col_vals("SITUAÇÃO")]).most_common():
    print(f"  {k:30s} {v}")

# Cobertura de coordenadas
def tem_coord(lat_col, lon_col):
    li, loi = idx_of(lat_col), idx_of(lon_col)
    n = 0
    for r in data:
        lat = r[li] if li is not None else None
        lon = r[loi] if loi is not None else None
        if lat not in (None, "") and lon not in (None, ""):
            n += 1
    return n

print("\n=== COBERTURA DE COORDENADAS ===")
print(f"  Barragem (orig):      {tem_coord('LATITUDE BARRAGEM','LONGITUDE BARRAGEM')}")
print(f"  Barragem (.KMZ):      {tem_coord('LAT. BARRAGEM .KMZ','LON. BARRAGEM .KMZ')}")
print(f"  Casa de força (orig): {tem_coord('LATITUDE CASA FORÇA','LONGITUDE CASA FORÇA')}")
print(f"  Casa de força (.KMZ): {tem_coord('LAT. CASA FORÇA .KMZ','LON. CASA FORÇA .KMZ')}")

wb.close()
