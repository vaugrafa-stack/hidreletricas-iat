"""
Monta o rastreador de conferência geoespacial (1 linha por empreendimento).

Lê o .xlsm em modo SOMENTE-LEITURA (não altera o original) e gera
data/processed/conferencia_rastreador.csv com:
  - dados de identificação + coordenadas originais e .KMZ (verbatim)
  - pontos de navegação normalizados (para abrir no Earth)
  - colunas VAZIAS a preencher na revisão visual (barragem + casa de força)

Se o rastreador já existir, PRESERVA o progresso já revisado (merge por id_emp).
"""
import os
import openpyxl
import pandas as pd
import yaml

import sys
sys.path.insert(0, os.path.dirname(__file__))
from geo_utils import safe_float, coords_validas, PR_BOUNDS

RASTREADOR = "data/processed/conferencia_rastreador.csv"

# Colunas que o revisor preenche (preservadas em merges futuros)
COLS_REVISAO = [
    "lat_barragem_conf", "lon_barragem_conf", "status_barragem_conf",
    "lat_casaforca_conf", "lon_casaforca_conf", "status_casaforca_conf",
    "grau_confianca", "fase_verificada", "obs_conferencia",
    "data_conferencia", "origem_conferencia", "revisado",
]


def chave_local(nav_lat, nav_lon, nome):
    """Chave da localização física: coord da barragem (~11 m) ou, sem coord, o nome."""
    la, lo = safe_float(nav_lat), safe_float(nav_lon)
    if la is not None and lo is not None:
        return f"{round(la, 4)},{round(lo, 4)}"
    return "NOME:" + str(nome).strip().upper()


def normaliza_ponto(lat_raw, lon_raw):
    """Devolve (lat, lon) plausível dentro do PR, aplicando heurísticas leves.
    NÃO altera o dado armazenado original — serve só para navegação no Earth."""
    lat = safe_float(lat_raw)
    lon = safe_float(lon_raw)
    if lat is None or lon is None:
        return None, None
    # já válido
    if coords_validas(lat, lon):
        return lat, lon
    # heurísticas: valor sem ponto decimal, lat/lon trocados, escala /1e5
    cand = [
        (lat, lon),
        (lat / 1e6, lon / 1e6),   # ex.: -25871039 -> -25.871039
        (lon, lat),               # lat/lon trocados
        (lat / 1e5, lon / 1e5),
    ]
    for la, lo in cand:
        if coords_validas(la, lo):
            return la, lo
    return None, None


def fase_provavel(situacao) -> str:
    if situacao is None or (isinstance(situacao, float)) or not str(situacao).strip():
        return "Indefinido"
    s = str(situacao).strip().upper()
    if s in ("DEFERIDO", "VIGENTE", "REGULAR"):
        return "Licenciado (verificar operação/instalação)"
    if s in ("PROTOCOLADO", "EM ANALISE", "EM ANÁLISE", "ANALISE REGIONAL"):
        return "Fase prévia (pode não estar construído)"
    if s in ("ARQUIVADO", "INDEFERIDO", "CANCELADO", "DEVOLVIDO - SGA", "SUSPENSO"):
        return "Não licenciado (provavelmente sem obra)"
    if s == "IBAMA":
        return "Licenciamento federal (IBAMA)"
    return "Indefinido"


def carregar_xlsm():
    with open("config.yaml", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    path = cfg["excel"]["local_path"]
    wb = openpyxl.load_workbook(path, read_only=True, keep_vba=True, data_only=True)
    ws = None
    for s in wb.sheetnames:
        if s.lower().startswith("central de processos hidrel"):
            ws = wb[s]
            break
    if ws is None:
        ws = wb.worksheets[0]

    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    last = max((i for i, v in enumerate(header) if v not in (None, "")), default=0) + 1
    header = [str(h).strip() if h is not None else f"_c{i}" for i, h in enumerate(header[:last])]

    data = [r[:last] for r in rows if any(v is not None for v in r[:last])]
    wb.close()
    return pd.DataFrame(data, columns=header)


def main():
    df = carregar_xlsm()

    def col(nome):
        return df[nome] if nome in df.columns else pd.Series([None] * len(df))

    out = pd.DataFrame()
    out["id_emp"] = range(1, len(df) + 1)
    out["codigo"] = col("CÓDIGO")
    out["protocolo"] = col("PROTOCOLO")
    out["empreendimento"] = col("NOME")
    out["tipo"] = col("TIPO")
    out["situacao"] = col("SITUAÇÃO")
    out["tipo_licenca"] = col("TIPOS DE LICENÇA")
    out["potencia_mw"] = col("POT SOLIC (MW)")
    out["rio"] = col("RIO")
    out["municipio"] = col("MUNICÍPIOS AFETADOS")

    # Coordenadas verbatim do .xlsm
    out["lat_barragem_orig"] = col("LATITUDE BARRAGEM")
    out["lon_barragem_orig"] = col("LONGITUDE BARRAGEM")
    out["lat_barragem_kmz"] = col("LAT. BARRAGEM .KMZ")
    out["lon_barragem_kmz"] = col("LON. BARRAGEM .KMZ")
    out["lat_casaforca_orig"] = col("LATITUDE CASA FORÇA")
    out["lon_casaforca_orig"] = col("LONGITUDE CASA FORÇA")

    # Pontos de navegação (normalizados; prioriza original, cai p/ KMZ)
    nav_bar_lat, nav_bar_lon, nav_cf_lat, nav_cf_lon = [], [], [], []
    for _, r in out.iterrows():
        la, lo = normaliza_ponto(r["lat_barragem_orig"], r["lon_barragem_orig"])
        if la is None:
            la, lo = normaliza_ponto(r["lat_barragem_kmz"], r["lon_barragem_kmz"])
        nav_bar_lat.append(la); nav_bar_lon.append(lo)
        cla, clo = normaliza_ponto(r["lat_casaforca_orig"], r["lon_casaforca_orig"])
        nav_cf_lat.append(cla); nav_cf_lon.append(clo)
    out["nav_lat_barragem"] = nav_bar_lat
    out["nav_lon_barragem"] = nav_bar_lon
    out["nav_lat_casaforca"] = nav_cf_lat
    out["nav_lon_casaforca"] = nav_cf_lon

    out["fase_provavel"] = out["situacao"].apply(fase_provavel)

    # Chave de localização física (para deduplicar processos da mesma usina)
    out["chave_local"] = [
        chave_local(r["nav_lat_barragem"], r["nav_lon_barragem"], r["empreendimento"])
        for _, r in out.iterrows()
    ]

    # Colunas de revisão (vazias)
    for c in COLS_REVISAO:
        out[c] = ""
    out["revisado"] = "não"

    # Merge com progresso anterior, se existir
    if os.path.exists(RASTREADOR):
        ant = pd.read_csv(RASTREADOR, encoding="utf-8-sig", dtype=str)
        if "id_emp" in ant.columns:
            ant = ant.set_index("id_emp")
            for _, r in out.iterrows():
                k = str(r["id_emp"])
                if k in ant.index and str(ant.loc[k].get("revisado", "não")).lower() == "sim":
                    for c in COLS_REVISAO:
                        if c in ant.columns:
                            out.loc[out["id_emp"] == r["id_emp"], c] = ant.loc[k].get(c, "")
            print("Progresso anterior preservado.")

    os.makedirs("data/processed", exist_ok=True)
    out.to_csv(RASTREADOR, index=False, encoding="utf-8-sig")

    # Resumo
    n = len(out)
    com_bar = out["nav_lat_barragem"].notna().sum()
    com_cf = out["nav_lat_casaforca"].notna().sum()
    print(f"Rastreador: {RASTREADOR}")
    print(f"  Registros de processo: {n}")
    print(f"  Localizações físicas únicas (a revisar): {out['chave_local'].nunique()}")
    print(f"  Com ponto navegável de BARRAGEM:      {com_bar}")
    print(f"  Com ponto navegável de CASA DE FORÇA: {com_cf}")
    print(f"  Sem casa de força (achar via conduto): {n - com_cf}")
    print("\n  Por fase provável:")
    print(out["fase_provavel"].value_counts().to_string())


if __name__ == "__main__":
    main()
