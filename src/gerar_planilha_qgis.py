"""Gera uma CÓPIA da planilha original com a coluna 'ABRIR NO QGIS'.

O .xlsm é um pacote ZIP. Em vez de reabrir no Excel (que recusa este arquivo) ou
no openpyxl (que descartaria os 12 gráficos), editamos CIRURGICAMENTE apenas o
XML da aba principal — todo o resto (gráficos, macros, tabelas, demais abas) é
copiado byte-a-byte. O original NÃO é alterado.

A nova coluna é uma fórmula HYPERLINK por linha (qgis://lat,lon), montada a
partir das colunas LATITUDE/LONGITUDE BARRAGEM. Por ser fórmula, atualiza-se
sozinha quando você corrigir as coordenadas.

Uso:  python src/gerar_planilha_qgis.py
"""
import os
import re
import sys
import shutil
import zipfile
from pathlib import Path

import yaml
import openpyxl

BASE = Path(__file__).resolve().parent.parent


def col_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def main():
    cfg = yaml.safe_load(open(BASE / "config.yaml", encoding="utf-8"))
    src = os.path.abspath(cfg["excel"]["local_path"])
    if not os.path.exists(src):
        print(f"[ERRO] Original não encontrado: {src}")
        return 1
    out = os.path.join(os.path.dirname(src), Path(src).stem + "_COM_QGIS.xlsm")

    # ── 1. Ler estrutura com openpyxl (cabeçalhos, colunas, última linha) ──
    wb = openpyxl.load_workbook(src, read_only=True, keep_vba=True, data_only=True)
    ws = wb.worksheets[0]
    sheet_title = ws.title
    header = {}
    last_header_col = 0
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None:
            header[str(cell.value).strip().upper()] = cell.column
            last_header_col = max(last_header_col, cell.column)

    def find_col(*names):
        for n in names:
            if n.upper() in header:
                return header[n.upper()]
        return None

    lat_c = find_col("LATITUDE BARRAGEM")
    lon_c = find_col("LONGITUDE BARRAGEM")
    if not lat_c or not lon_c:
        print("[ERRO] Não achei LATITUDE/LONGITUDE BARRAGEM. Cabeçalhos:", list(header)[:40])
        wb.close()
        return 1

    # última linha com PROTOCOLO (col 2) preenchido
    last_row = 1
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        if row[0].value is not None:
            last_row = row[0].row
    wb.close()

    new_c = last_header_col + 1
    lat, lon, newcol = col_letter(lat_c), col_letter(lon_c), col_letter(new_c)
    print(f"Aba: {sheet_title} | lat={lat} lon={lon} | nova coluna={newcol} | linhas 2..{last_row}")

    # ── 2. Descobrir o arquivo XML da 1ª aba dentro do ZIP ──
    shutil.copy2(src, out)
    with zipfile.ZipFile(out) as z:
        wbxml = z.read("xl/workbook.xml").decode("utf-8")
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    m_sheet = re.search(r"<sheet\b[^>]*>", wbxml)
    rid = re.search(r'r:id="([^"]+)"', m_sheet.group(0)).group(1)
    m_rel = re.search(r'<Relationship[^>]*Id="%s"[^>]*?Target="([^"]+)"' % re.escape(rid), rels) \
        or re.search(r'<Relationship[^>]*Target="([^"]+)"[^>]*?Id="%s"' % re.escape(rid), rels)
    target = m_rel.group(1).lstrip("/")
    sheet_path = target if target.startswith("xl/") else "xl/" + target

    # ── 3. Editar o XML da aba: cabeçalho + fórmulas + dimension ──
    with zipfile.ZipFile(out) as z:
        sheet_xml = z.read(sheet_path).decode("utf-8")

    f_raw = (f'IF({lat}{{n}}="","",HYPERLINK("qgis://"'
             f'&SUBSTITUTE({lat}{{n}}&"",",",".")&","'
             f'&SUBSTITUTE({lon}{{n}}&"",",",".")'
             f',"Abrir no QGIS"))')

    def header_cell():
        return f'<c r="{newcol}1" t="inlineStr"><is><t>ABRIR NO QGIS</t></is></c>'

    def formula_cell(n):
        f = f_raw.format(n=n).replace("&", "&amp;")
        return f'<c r="{newcol}{n}" t="str"><f>{f}</f></c>'

    def repl(mt):
        rn = int(mt.group(2))
        if rn == 1:
            return mt.group(1) + mt.group(3) + header_cell() + mt.group(4)
        if 2 <= rn <= last_row:
            return mt.group(1) + mt.group(3) + formula_cell(rn) + mt.group(4)
        return mt.group(0)

    sheet_xml = re.sub(r'(<row r="(\d+)"[^>]*>)(.*?)(</row>)', repl, sheet_xml, flags=re.DOTALL)
    sheet_xml = re.sub(r'<dimension ref="[^"]*"\s*/>',
                       f'<dimension ref="A1:{newcol}{last_row}"/>', sheet_xml, count=1)

    # ── 4. Forçar recálculo ao abrir (workbook.xml) ──
    if "<calcPr" in wbxml:
        wbxml2 = re.sub(r'<calcPr\b', '<calcPr fullCalcOnLoad="1"', wbxml, count=1)
        wbxml2 = wbxml2.replace(' fullCalcOnLoad="1" fullCalcOnLoad=', ' fullCalcOnLoad=')
    else:
        wbxml2 = wbxml.replace("</workbook>", '<calcPr fullCalcOnLoad="1"/></workbook>')

    # ── 5. Reescrever o ZIP preservando todo o resto byte-a-byte ──
    tmp_out = out + ".tmp"
    with zipfile.ZipFile(out) as zin, zipfile.ZipFile(tmp_out, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_path:
                data = sheet_xml.encode("utf-8")
            elif item.filename == "xl/workbook.xml":
                data = wbxml2.encode("utf-8")
            zout.writestr(item, data)
    os.replace(tmp_out, out)

    print(f"[OK] Planilha gerada: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
